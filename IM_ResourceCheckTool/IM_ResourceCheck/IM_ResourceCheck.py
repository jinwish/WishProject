# -*- coding: utf-8 -*-
#!/usr/bin/env python

# 한글 문자열 수집을 위해 필요함
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

import unreal_engine as ue
import openpyxl
from unreal_engine import SWindow, SButton, STextBlock, SHorizontalBox, SVerticalBox, SGridPanel, SScrollBox, SBorder, SEditableTextBox, SCheckBox
from unreal_engine.enums import EHorizontalAlignment, EOrientation, EVerticalAlignment, EHorizontalAlignment
from unreal_engine.classes import DataTableFactory
from unreal_engine.structs import Margin, Key
from IM_ResourceCheck import IM_Slate as newslate
from IM_ResourceCheck import IM_SlateUtil as newutil
from IM_UtilClass import IM_UtilClass as textw

ctext = newutil.IMText()
arraylist = []
textarray = []

sort_bool = SCheckBox() # Sort 할것인가?
checkbox_bool = SCheckBox() # 매칭되지 않는 리스트만 출력
table_path = None 
bntlists = [] #버튼 리스트 저장
tablename = ''

# 버튼 리스트 클래스 (버튼외 row 값을 클래스로 만들어야 리스트 갱신이 가능하다.)
class SetButtontlist:
	def __init__(self,num):
		self.num_bnt = num
		self.set_bnt = SButton(text= '       리스트확인        ', on_clicked=lambda: fn_bntindex(self.num_bnt))

# 버튼의 인덱스 값을 부여 한다.
def fn_bntindex(num):
	inx = num
	fn_loadtablesetting(table_path.get_text(),inx)
	#ue.log('리스트 인덱스 값 확인 : ' + str(inx))
	#ue.log(num)

class SetWindow:
	def __init__(self):
		self.window = SWindow(client_size=(680, 530), title='Resource List Check', sizing_rule=0)
		self.hori_0 = SHorizontalBox()
		self.hori_0_0 = SHorizontalBox()
		self.hori_1_0 = SHorizontalBox()
		self.hori_2 = SHorizontalBox()
		self.hori_2_0 = SHorizontalBox()
		#self.hori_3_0 = SHorizontalBox()
		self.vert_0 = SVerticalBox()
		#self.vert_2 = SVerticalBox()
		self.main_vertical = SVerticalBox()

		self.scroll = SScrollBox(orientation=EOrientation.Orient_Vertical)
		self.chk_grid = SGridPanel()

		top_vertical = (0,0,0,0)
		mid_vertical = (0,-400,0,-390)
		bot_vertical = (0,390,0,0)
		#zero_padding = (0,0,0,0)

		self.newindex = []
		self.newint = []

		self.txt_01 = STextBlock(text='엑셀 테이블 경로 : ')
		self.etb = SEditableTextBox().assign('table_path')

		self.hori_0.add_slot(self.txt_01, fill_width=0.1)
		self.hori_0.add_slot(self.etb, fill_width=0.4)
		self.hori_0.add_slot(SButton(text='...', fill_width=0.1, on_clicked=lambda: self.fn_sp()), auto_width=True)
		self.hori_0.add_slot(SButton(text='@', fill_width=0.1, on_clicked=lambda: self.fn_test()), auto_width=True)
		self.hori_0_0.add_slot(SBorder()(self.hori_0), padding=top_vertical, v_align=EVerticalAlignment.VAlign_Top)

		self.hori_1_0.add_slot(SBorder()(self.vert_0), padding=mid_vertical ,v_align=EVerticalAlignment.VAlign_Fill)
		
		self.hori_2.add_slot(SButton(text='선택리스트 저장-TXT', h_align=2,v_align=2, fill_width=0.1, on_clicked=lambda: fn_sel()))
		self.hori_2.add_slot(SButton(text='저장리스트 비교-TXT', h_align=2,v_align=2, fill_width=0.1, on_clicked=lambda: fn_readtext()))
		self.hori_2.add_slot(STextBlock(text='  매칭 리스트 정렬 '), auto_width=True, v_align=EVerticalAlignment.VAlign_Center)
		self.hori_2.add_slot(sort_bool, auto_width=True)
		self.hori_2.add_slot(STextBlock(text='  리스트에 없는 에셋만 표시 '), auto_width=True, v_align=EVerticalAlignment.VAlign_Center)
		self.hori_2.add_slot(checkbox_bool, auto_width=True)
		self.hori_2_0.add_slot(SBorder()(self.hori_2), padding=bot_vertical ,v_align=EVerticalAlignment.VAlign_Fill)

		# 초기설정파일 셋팅
		if fn_OpenDefaultSetting() != None:
			table_path.set_text(fn_OpenDefaultSetting())
			#테이블 리스트 함수 - 설정 경로 바탕으로 리스트를 작성 합니다
			if self.fn_resettable() != None:
				self.fn_resettable()
		else:
			ue.log('설정 파일이 없습니다')

		self.main_vertical.add_slot(self.hori_0_0)
		self.main_vertical.add_slot(SBorder()(self.scroll),padding=mid_vertical, v_align=EVerticalAlignment.VAlign_Fill)
		self.main_vertical.add_slot(self.hori_2_0)

		self.window.set_content(SBorder()(self.main_vertical))

	# 테이블 버튼 리스트 생성
	def fn_resettable(self):
		
		table_padding = (10, 5, 10, 5)
		btn_padding = (4, 4, 4, 4)

		# 기본 리스트 설정
		self.chk_grid.add_slot(SBorder(padding=table_padding,h_align=2,v_align=2)(STextBlock(text=ctext.tableHeader_01)), column=0, row=0)
		self.chk_grid.add_slot(SBorder(padding=table_padding,h_align=2,v_align=2)(STextBlock(text=ctext.listHeader_01)), column=1, row=0)
		self.chk_grid.add_slot(SBorder(padding=table_padding,h_align=2,v_align=2)(STextBlock(text=ctext.listHeader_02)), column=2, row=0)
		self.chk_grid.add_slot(SBorder(padding=table_padding,h_align=2,v_align=2)(STextBlock(text=ctext.listHeader_03)), column=3, row=0)
		self.chk_grid.add_slot(SBorder(padding=table_padding,h_align=2,v_align=2)(STextBlock(text=ctext.listHeader_04)), column=4, row=0)
		
		# 테이블에서 엑셀 경로 가져오기
		listxx = fn_Dafultxlsx(table_path.get_text())
		if listxx == None:
			return None
		#ue.log(listxx)

		# 리스트 생성
		for i in range(0,len(listxx)):
			bntlists.append(SetButtontlist(i))

		i = 1
		x = 0

		for ttable in listxx:
			d = i
			name = ttable[0] #'엑셀파일 이름'
			sname = ttable[1] #'시트 이름'
			cname = ttable[2] #'컬럼 이름'
			uesasset = ttable[3] #'리소스 경로'
			
			self.chk_grid.add_slot(SBorder(padding=table_padding,h_align=2,v_align=2)(SEditableTextBox(text=str(i))), column=0, row=i)
			self.chk_grid.add_slot(SBorder(padding=table_padding,h_align=2,v_align=2)(STextBlock(text=name)), column=1, row=i)
			self.chk_grid.add_slot(SBorder(padding=table_padding,h_align=2,v_align=2)(STextBlock(text=sname)), column=2, row=i)
			self.chk_grid.add_slot(SBorder(padding=table_padding,h_align=2,v_align=2)(STextBlock(text=cname)), column=3, row=i)
			# 버튼 리스트 ADD
			self.chk_grid.add_slot(SBorder(padding=btn_padding,v_align=2)(bntlists[i-1].set_bnt), column=4, row=i)			
			i += 1

		self.scroll.add_slot(self.chk_grid, auto_width = True, auto_hight = True)

	def fn_test(self):
		ue.log('설정테스트')

	# 테이블 경로 - 가져오기
	def fn_sp(self):
		try:
			st = fn_SaveDefaultSetting()
			table_path.set_text(st)
			if self.fn_resettable() != None:
				self.fn_resettable()
		except:
			ue.log('엑셀 기본 경로 설정을 취소 하였습니다')
	
	# 엑셀 테이블 읽기 위한 필요한 정보 (삭제 해야함
	def fn_tt(self,path,tint):
		datacolumn = fn_Dafultxlsx(path)
		#ue.log(len(datacolumn))
		#ue.log(tint)
		filepath = path + '/' + datacolumn[tint][0] + '.xlsx'
		sheetname = datacolumn[tint][1]
		findcolumn = datacolumn[tint][2]
		addstring = datacolumn[tint][3]
		fn_Loadxlsx(filepath, sheetname, findcolumn, addstring)

# 엑셀 테이블 읽기 위한 필요한 정보 (전역 함수로 변경)
def fn_loadtablesetting(path,tint):
	datacolumn = fn_Dafultxlsx(path)
	filepath = path + '/' + datacolumn[tint][0] + '.xlsx'
	sheetname = datacolumn[tint][1]
	findcolumn = datacolumn[tint][2]
	addstring = datacolumn[tint][3]
	fn_Loadxlsx(filepath, sheetname, findcolumn, addstring)

# 리소스 경로 저장된 엑셀파일 읽기
def fn_Dafultxlsx(datatablepath):
	xslxpath = datatablepath + '/TableResourcePath.xlsx'
	stringarray = []
	sstring = []
	sname = 'TableResourcePath'
	try:
		load_wb = openpyxl.load_workbook(xslxpath, data_only=True)
	except:
		ue.log('TableResourcePath.xlsx 엑셀파일을 찾지 못하였습니다.')
		return None
	

	load_ws = load_wb[sname]

	indexY = 2
	indexX = 1
	while True:
		#ue.log(load_ws.cell(indexY,indexX).value)
		sstring.append(load_ws.cell(indexY,indexX).value)
		indexX += 1
		if not load_ws.cell(indexY,indexX).value:
			stringarray.append(sstring)
			indexY += 1
			indexX = 1
			sstring = []
			if not load_ws.cell(indexY,indexX).value:
				break
	return stringarray

# 엑셀파일 불러오기
def fn_Loadxlsx(filepath,sheetname,findcolumn,addstring):
	#ue.log('데이타 확인용 : ' + filepath + ' ' + sheetname + ' ' + findcolumn + ' ' + addstring )
	tablename = ' - ' + filepath + ' - ' + sheetname + ' - ' + findcolumn + '\n'
	xslxpath = filepath #'E:/SkillEffectParam.xlsx' #filepath
	sheetListval = []
	column = 1
	row = 1
	sname = sheetname #'Buff' #sheetname #시트명 #TableResourcePath
	fcolumn = findcolumn #'fx_path|c' #'icon_path|c' #findcolumn #컬럼명
	astring = '/Game/' + addstring + '/' #addstring
	addcomma = False
	try:
		load_wb = openpyxl.load_workbook(xslxpath, data_only=True)
		# 원하는 시트 / 원하는 sell 읽기
		load_ws = load_wb[sname]
	except:
		ue.log('리스트에 있는 엑셀파일을 찾지 못하였습니다.')
		return 

	# 특정 열 수집
	index = 1
	column = index
	while True:
		column = index
		if load_ws.cell(1,index).value == fcolumn:break
		index += 1
		if not load_ws.cell(1,index).value: break
	index = 2
	row = index
	while True:
		addcomma = False
		row = index
		if load_ws.cell(index,column).value != '*':
			xc = load_ws.cell(index,column).value
			xs = xc.split('/')
			ast = ''
			
			#ue.log('. 확인해보자' + str(len(xc.split('.'))))

			# 엑셀테이블에 Dot 이 있는지 검사
			if len(xc.split('.')) == 2:addcomma = False
			else:addcomma = True

			# 문자열 재조합
			if addcomma == True:
				ast = astring + xc + '.' + xs[len(xs) - 1]
			else:
				ast = astring + xc
			sheetListval.append(ast)
		index += 1
		if not load_ws.cell(index,column).value: break
	
	#for sh in sheetListval:
	#	ue.log('경로완성 : ' + sh)

	#if sort_bool == True:
	if sort_bool.is_checked():
		sheetListval.sort()

	#테이블
	fn_setTable(sheetListval,tablename) 
	
# 불필요한 문자열 제외한다 - 사용안함
def fn_stringparsing(stringarray):
	finalall = []
	for strings in stringarray:
		if strings != '*': #해당 문제는 수집하지 않는다
			strsplit = strings.split('/')
			final = '/Game/' + strings + '.' + strsplit[len(strsplit) - 1]
			#ue.log(final)
			finalall.append(final)
	return finalall

#테이블 작성	
def fn_setTable(stringlist, tablename):
	newarraylist = []
	if len(stringlist) == 0: return
	#버프 아이콘 경로 변경
	
	arraylist = stringlist
	#ue.log('정말 진실인가? : ' + str(cb_bool))
	for ttable in arraylist:
		if ue.find_asset(ttable) == None:
			#ue.log('정말 없다!')
			className = '?????' #ue.find_asset(ttable).get_class().get_name() # #asset.className
			path = ttable
			assetname = ttable.split('.')[1]
			uesasset = '없음'
			addall = className + ',' + path + ',' + assetname + ',' + uesasset
			newarraylist.append(addall)
		else:
			#ue.log('정말 있다!')
			if checkbox_bool.is_checked() == False:
				className = ue.find_asset(ttable).get_class().get_name()
				path = ttable
				assetname = ttable.split('.')[1]
				uesasset = '있음'
				addall = className + ',' + path + ',' + assetname + ',' + uesasset
				newarraylist.append(addall)
	ue.log('제대로나옴? : ' + tablename)
	#테이블에 작성
	newslate.ClassTable(ctext, newarraylist, tablename)

# 선택한 에셋 경로 저장
def fn_sel():
	arraylist = []
	selbp = ue.get_selected_assets()
	for i in selbp:
		arraylist.append(i.get_path_name())
	fn_writetext(arraylist)

def fn_writetext(textarray):
	try:
		newpath = []
		filename = ue.save_file_dialog('Save a Text file', '', '', 'TXT|*.txt')[0]
		for i in textarray:
			newpath.append(i + '\n')
		textw.TextFileWrite(newpath,filename)
	except:
		ue.log('### 저장을 취소하였습니다###')

# 선택한 에셋 경로 리스트로 현재 프로젝트를 비교
def fn_readtext():
	try:
		filename = ue.open_file_dialog('Load a Text file', '', '', 'TXT|*.txt')[0]
		textarray = []
		textarray = textw.TextFileRead(filename)
		ue.log('### 불러오기 성공 하였습니다 ###')
		openl = fn_setTable(textarray.Load_list)
	except:
		ue.log('### 불러오기을 취소하였습니다###')

# 체크툴 기본 값을 저장 한다
def fn_SaveDefaultSetting():
	try:
		openpath = ue.open_directory_dialog('')
		dt_factory = DataTableFactory()
		dt_factory.Struct = Key
		key = Key()
		key.KeyName = openpath
		dt = dt_factory.factory_create_new('/Game/Assets/ZExperiment/Scripts/IM_ResourceCheck/DefaultSetting')
		dt.data_table_add_row("{0}".format('TablePath_' + str(0)), key)
		dt.save_package()
		return openpath
	except:
		return ''

# 체크툴 기본 값을 불러온다
def fn_OpenDefaultSetting():
	try:
		datat = ue.get_asset('/Game/Assets/ZExperiment/Scripts/IM_ResourceCheck/DefaultSetting' + '.DefaultSetting')
		loadsetting = datat.data_table_find_row('TablePath_0').KeyName
		#ue.log(loadsetting)
		#ue.log('설정파일있음')
		return loadsetting
	except:
		ue.log('설정파일없음')
		return None

if __name__ == '__main__':
    window = SetWindow()